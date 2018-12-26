# _*_ coding:utf-8 _*_
from openpyxl import load_workbook
import bean.Paper as p
import os

NUMBER = 1
TITLE = 2
AUTHOR = 3
PUBLISH = 4
URL = 5
ABSTRACT = 6
CITATION_NAME = 7
CITATION_URL = 8
REFERENCE_NAME = 9
REFERENCE_URL = 10
CITATION_NUM = 11

list_cita_title = []
list_cita_url = []
list_refer_title = []
list_refer_url = []


def write_file(filepath):
    wb = load_workbook("database.xlsx")
    sheet = wb.worksheets[0]

    # 载入现有的database最后一行的下一行
    length = len(sheet["B"]) + 1
    pathDir = os.listdir(filepath)
    # 激活sheet
    wb1 = wb.active

    # 传入的number
    count = 274
    for s in pathDir:
        newDir = os.path.join(filepath, s)
        if os.path.isfile(newDir):
            if os.path.splitext(newDir)[1] == ".txt":
                p = readFile(newDir, count)

                wb1.cell(length, NUMBER, p.number)
                wb1.cell(length, TITLE, p.title)
                wb1.cell(length, AUTHOR, p.authors)
                wb1.cell(length, PUBLISH, p.published_in)
                wb1.cell(length, URL, p.url)
                wb1.cell(length, ABSTRACT, p.abstract)
                wb1.cell(length, CITATION_NAME, str(p.citations_name)[1:-1])
                wb1.cell(length, CITATION_URL, str(p.citations_url)[1:-1])
                wb1.cell(length, REFERENCE_NAME, str(p.references_name)[1:-1])
                wb1.cell(length, REFERENCE_URL, str(p.references_url)[1:-1])
                wb1.cell(length, CITATION_NUM, p.citation_number)

                count += 1
                length += 1
    wb.save("database.xlsx")

def readFile(filepath, num):
    """
        set时存的是 目前获取的该paper最大的citation number
        get时放的是 citation list的长度 和 number of citation 中的最大值
    """
    # 打开传进来的路径

    f1 = open(filepath, "r",encoding='UTF-8')
    # 读取所有行
    lines = f1.readlines()
    # 去掉已经被读取的前7行
    lines_count = len(lines) - 7

    name = str(lines[0])[6:]
    print(num,name)

    url = str(lines[1])[5:]
    publish_in = str(lines[2])[11:]
    author = str(lines[3])[9:-1]
    abstract = str(lines[4])[10:]
    # try:
    #     citation_num = int(str(lines[5])[18:])
    # except:
    #     print(lines[5])
    citation_num = int(str(lines[5])[18:])
    list_cita_title.clear()
    list_cita_url.clear()
    list_refer_title.clear()
    list_refer_url.clear()

    # 计算citation数目
    count = 0

    # 读取citation
    for n in range(lines_count):
        line = lines[n + 7]
        if line[:10] != "References":
            try:
                index_ = line.index("http:")
                name = line[:index_]
                url = line[index_:]
                # print(name)
                list_cita_title.append(name)
                list_cita_url.append(url)
                count += 1
            except :
                # print(line)
                # print(type(line))
                pass
        else:
            break

    # 读取reference
    lines_count -= count
    add_num = 8 + count
    for n in range(lines_count-1):
        line = lines[n + add_num]
        try:
            index_ = line.index("http:")
            name = line[:index_]
            url = line[index_:]
            list_refer_title.append(name)
            list_refer_url.append(url)
        except:
            print(filepath)
            print(line)
    return p.PaperBean(number=num, title=name, authors=author, published_in=publish_in, url=url,
                       abstract=abstract, citations_name=list_cita_title, citations_url=list_cita_url,
                       references_name=list_refer_title, reference_url=list_refer_url, citation_number=citation_num)

if __name__ == '__main__':
    # 填入txt所在的路径
    write_file("E:\ChormeDownload\S_Paper_Ranking\dao\CARP")
