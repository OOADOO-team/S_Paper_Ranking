# re.findall(r"\d+\.?\d*", value_of_block)
# _*_ coding:utf-8 _*_
import os

from openpyxl import load_workbook

NAME = 1
IF = 2
name = []
if_factor = []

def write_publish(filepath):
    wb = load_workbook("publish.xlsx")
    sheet = wb.worksheets[0]

    # 载入现有的database最后一行的下一行
    length = len(sheet["B"]) + 1
    pathDir = os.listdir(filepath)
    # 激活sheet
    wb1 = wb.active

    # 传入的number
    count = 0
    for s in pathDir:
        newDir = os.path.join(filepath, s)
        if os.path.isfile(newDir):
            if os.path.splitext(newDir)[1] == ".txt":
                length = readFile(newDir, count)
                for i in range(length):
                    wb1.cell(i+1, NAME, name[i])
                    wb1.cell(i+1, IF, if_factor[i])

                count += 1
                length += 1
    wb.save("database.xlsx")


def readFile(filepath, num):

    # 打开传进来的路径
    name.clear()
    if_factor.clear()
    f1 = open(filepath, "r", encoding='UTF-8')
    # 读取所有行
    lines = f1.readlines()
    length = len(lines)
    for line in lines:
        value = float(line.split(" ")[-1])
        publish_name = line.replace(str(value), "")
        name.append(publish_name)
        if_factor.append(value)
    return length

if __name__ == '__main__':
    # 填入txt所在的路径
    write_publish("E:\ChormeDownload\S_Paper_Ranking\dao")
