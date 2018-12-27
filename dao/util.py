
from openpyxl import load_workbook
from numpy import *
from dao.publishedDao import *

def convert():
    pe = load_workbook("E:\ChormeDownload\S_Paper_Ranking\dao\publish.xlsx")  # 默认可读写，若有需要可以指定write_only和read_only为True
    sheet = pe.worksheets[0]
    print("load database is correct")
    length = len(sheet["B"])
    publish_name = []
    publish_if = []
    for i in range(1, length + 1):
        pb = publishedBean(published_name=str(sheet.cell(i, 1).value),
                           IF_value=float(sheet.cell(i, 2).value))
        insert_DB(pb)

if __name__ == '__main__':

    convert()