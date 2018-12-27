
import re
from openpyxl import load_workbook
import bean.Paper as p
import os

NUMBER = 1
TITLE = 2
AUTHOR = 3
PUBLISH = 4
URL = 5
ABSTRACT = 6
CITATION_NAME = 7
CITATION_URL = 8
REFERENCE_NAME = 9
REFERENCE_URL = 10
CITATION_NUM = 11

def re_write(path):

    wb1 = load_workbook("database.xlsx")
    wb_write = wb1.active
    wb = load_workbook(path)
    sheet = wb.worksheets[0]
    length = len(sheet["B"])
    result = []
    # get each line of xlsx
    sentence = ""
    exist = set()
    count = 1
    for i in range(length):
        # get the value of B0 to Bn
        # print("read ",i)
        for j in range(2,12):
            sentence += str(sheet.cell(row=i + 2, column=j).value)
        judge = re.search("\\\\u", sentence)
        if judge is None:
            print(count)
            wb_write.cell(count, NUMBER, count)
            wb_write.cell(count, TITLE, sheet.cell(row=i + 2, column=2).value)
            wb_write.cell(count, AUTHOR, sheet.cell(row=i + 2, column=3).value)
            wb_write.cell(count, PUBLISH, sheet.cell(row=i + 2, column=4).value)
            wb_write.cell(count, URL, sheet.cell(row=i + 2, column=5).value)
            wb_write.cell(count, ABSTRACT, sheet.cell(row=i + 2, column=6).value)
            wb_write.cell(count, CITATION_NAME, sheet.cell(row=i + 2, column=7).value)
            wb_write.cell(count, CITATION_URL, sheet.cell(row=i + 2, column=8).value)
            wb_write.cell(count, REFERENCE_NAME, sheet.cell(row=i + 2, column=9).value)
            wb_write.cell(count, REFERENCE_URL, sheet.cell(row=i + 2, column=10).value)
            wb_write.cell(count, CITATION_NUM, sheet.cell(row=i + 2, column=11).value)
            count += 1
        else:
            # print(sentence)
            pass
    wb1.save("database.xlsx")



if __name__ == '__main__':
    re_write("E:\ChormeDownload\S_Paper_Ranking\dao\paper.xlsx")