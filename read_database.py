from openpyxl import load_workbook
import bean.Paper as p
import re
import rank as r


# ================================================================
def get_infomation(keyword, alpha):
    """
        通过获取用户输入于界面的keyword,搜索数据库中author和title是否有吻合的关键词
        如果有吻合的关键词，则利用PaperBean构造出paper
        将paper加入result中，最后返回result list
        支持模糊搜索

        :param keyword: 用户输入于界面的keyword
        :return result： paper list
    """
    if keyword is "":
        keyword = " "
    wb = load_workbook("dao\paper.xlsx")
    sheet = wb.worksheets[0]
    length = len(sheet["B"])
    result = []
    # get each line of xlsx
    exist = set()
    print('Reading!!!!!')
    for i in range(length):
        # get the value of B0 to Bn
        value_of_blockBn = str(sheet.cell(row=i + 2, column=2).value).upper()
        # print("value_of_blockBn",value_of_blockBn)
        value_of_blockCn = str(sheet.cell(row=i + 2, column=3).value).upper()
        # print("value_of_blockCn is ",value_of_blockCn)
        data_U = keyword.upper()
        # remove the useless word
        judge1 = re.search(data_U, value_of_blockBn)
        judge2 = re.search(data_U, value_of_blockCn)
        if judge1 is not None or judge2 is not None:
            if str(sheet.cell(row=i + 2, column=2).value) not in exist:
                # print(str(sheet.cell(row=i + 2, column=2).value))
                paper = p.PaperBean(number=i,
                                    title=str(sheet.cell(row=i + 2, column=2).value),
                                    authors=str(sheet.cell(row=i + 2, column=3).value),
                                    published_in=str(sheet.cell(row=i + 2, column=4).value),
                                    url=str(sheet.cell(row=i + 2, column=5).value),
                                    abstract=str(sheet.cell(row=i + 2, column=6).value),
                                    citations_name=list(str(sheet.cell(row=i + 2, column=7).value)[1:-1].split("', '")),
                                    citations_url=list(
                                        str(sheet.cell(row=i + 2, column=8).value).replace("'", "").replace(",",
                                                                                                            "").replace(
                                            "\\n", "").split()),
                                    references_name=list(str(sheet.cell(row=i + 2, column=9).value)[1:-1].split("', '")),
                                    references_url=list(
                                        str(sheet.cell(row=i + 2, column=10).value).replace("'", "").replace(",",
                                                                                                             "").replace(
                                            "\\n", "").split()),
                                    citations_number=int(sheet.cell(row=i + 2, column=11).value) if sheet.cell(row=i + 2,
                                                                                                              column=11).value else 0
                                    )
                result.append(paper)
                exist.add(str(sheet.cell(row=i + 2, column=2).value))
    result = r.rank_simple(result,alpha)
    print('Result return!!!!!')
    return result


if __name__ == '__main__':
    result = get_infomation(keyword="machine", alpha=100)
    # for item in result:
    #     print(item[1].title,item[1].authors, item[1].published_in)
    #     print(item.title)
    #     print()
