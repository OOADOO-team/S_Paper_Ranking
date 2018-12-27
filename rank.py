import bean.Paper as p
from openpyxl import load_workbook
import bean.Paper as p
import os
import re
import numpy as np
from numpy import *
import copy
import time
from dao.publishedDao import *

ITERATION_TIME = 10

def rank_simple(paperlist, alpha):


    # 读取publish的name和对应的IF值
    # pe = load_workbook("dao\publish.xlsx")  # 默认可读写，若有需要可以指定write_only和read_only为True
    # sheet = pe.worksheets[0]
    # print("load database is correct")
    # length = len(sheet["B"])
    # publish_name = []
    # publish_if = []
    # for i in range(1, length + 1):  # range默认从0开始，到后面参数的-1结束，而openpyxl都是从第一行第一列开始的，所以参数为1，maxC+1；意思就是遍历第一列到最后一列，
    #     publish_name.append(str(sheet.cell(i, 1).value).upper())
    #     # print(sheet.cell(i, 2).value)
    #     publish_if.append(float(sheet.cell(i, 2).value))
    #

    print("read publish is correct")
    # ==== 读取paper的citation的值，找到对应的publish并赋值。最后加权排序 =====

    total_list = []

    # 根据10/79.258得到的publish_value应该乘的weight
    weight = 0.12617
    for paper in paperlist:
        # print("跑paperlist呢")
        citation = paper.citations_number
        publish = paper.published_in
        valuep = 0
        valuec = find_citation_value(citation)
        flag = False

        # 找publish
        valuep = read_DB(publish)
        if valuep is not None:
            valuep *= weight
            flag = True
            break
        if flag==False:
            valuep = 0.189255
        alpha = int(alpha)
        total_value = valuep * alpha + (100 - alpha) * valuec

        total_list.append((paper, total_value))

    # 根据total_value对paper进行排序
    _list = sorted(total_list, key=lambda x: x[1])
    return_list = [item[0] for item in _list]
    for item in _list:
        print(item[0].title,item[1])
    return_list = paper_rank(return_list, ITERATION_TIME)
    return return_list


def find_citation_value(citation):
    if citation > 1000:
        return 10
    elif citation > 900:
        return 9
    elif citation > 800:
        return 8
    elif citation > 700:
        return 7
    elif citation > 600:
        return 6
    elif citation > 500:
        return 5
    elif citation > 400:
        return 4
    elif citation > 300:
        return 3
    elif citation > 200:
        return 2
    else:
        return 1


def degree(figure):
    '''return in_result, out_result of each node in the figure'''
    out_result = {node: 0
                 for node in figure}
    in_result = {node: 0
                 for node in figure}
    # print(figure[1])
    for i in figure:
        # print(type(i))
        for j in figure[i]:
            in_result[j] += 1
            out_result[i] += 1
    mean1 = average(list(in_result.values()))
    var1 = np.std(list(in_result.values()))
    mean2 = average(list(out_result.values()))
    var2 = np.std(list(out_result.values()))
    for i in figure:
        in_result[i] = (in_result[i]-mean1) / var1
        out_result[i] = (out_result[i]-mean2) /var2
    return in_result, out_result


def eigenvector_centrality(figure):
    '''return eigenvector_centrality of each node in figure'''
    # centrality = nx.eigenvector_centrality_numpy(figure)
    # import scipy as sp
    nodelist = list(figure)
    length = len(figure)
    index = dict(zip(nodelist, range(length)))

    M = np.zeros((length, length))
    for i in figure:
        for j in figure[i]:
            M[index[i]][index[j]] = 1
            M[index[j]][index[i]] = 1
    u = 1/length
    red_eig_vects = np.zeros(length)
    for i in range(length):
        red_eig_vects[i] = u
    red_eig_vects = mat(red_eig_vects).T
    for i in range(200):
        red_eig_vects = M * red_eig_vects
        red_eig_vects /= linalg.norm(red_eig_vects)
    return dict(zip(figure, red_eig_vects))


def paper_rank(figure, itera):
    temp = copy.deepcopy(figure)
    try:
        figure = range(itera)
        length = len(figure)
        out_d = np.zeros([length])
        nodelist = list(figure)
        index = dict(zip(nodelist, range(length)))
        for i in figure:
            for j in figure[i]:
                out_d[index[i]] += 1
        length = len(figure)
        nodelist = list(figure)
        out_mat = np.zeros((length, length))
        for i in range(length):
            out_mat[i][i] = out_d[i]
        out_mat[out_mat != 0] = 1 / out_mat[out_mat != 0]
        index = dict(zip(nodelist, range(length)))
        M = np.zeros((length, length))
        for i in figure:
            for j in figure[i]:
                M[index[i]][index[j]] = 1
        _P = out_d * M
        w_ = np.array(1 / length)
        w = mat(w_.repeat(length)).T
        d = np.zeros([length])
        for i in range(length):
            if out_d[i] == 0:
                d[i] = 1
        P = _P + w*d
        e = mat(np.ones(length)).T
        v = w_.repeat(length)
        G = 0.85 * P + 0.15 * e * v
        pai = v
        for i in range(itera):
            pai = pai*G
            pai = pai/np.linalg.norm(pai)
    except:
        pass
    result = temp
    return result